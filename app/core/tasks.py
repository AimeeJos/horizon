from __future__ import absolute_import, unicode_literals
from celery import shared_task
import time
import openpyxl
from openpyxl.styles import Font  # PatternFill
import os
from django.conf import settings
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from transformers import pipeline
from core.models import Sentences, Categories


def write_to_excel_file(self, finaldatas):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    font = Font(bold=True)
    fields = list(finaldatas[0].keys())

    for i in range(len(fields)):
        cell = worksheet.cell(row=1, column=(i + 1), value=fields[i])
        cell.font = font

    c = 2
    values = []
    total_steps = len(finaldatas)
    print("=== total: ", total_steps)
    for data in finaldatas:
        values = list(data.values())

        # !progress bar
        self.update_state(state="PROGRESS", meta={"current": c, "total": total_steps})
        percentage_complete = int(((c - 1) / total_steps) * 100)
        task_progress = {
            "status": percentage_complete,
            "taskid": self.request.id,
        }
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            self.request.id,  # Use the task ID as the channel group name
            {"type": "task_progress", "percentage_complete": task_progress},
        )
        # !########## progres bar end #############

        for i in range(len(fields)):
            worksheet.cell(row=c, column=(i + 1), value=str(values[i]))
        c = c + 1

    # Create a directory to save the Excel file
    taskid = self.request.id
    directory = os.path.join(settings.MEDIA_ROOT, "downloads")
    # print(f"____directory___{directory} exists?: {os.path.exists(directory)}")
    if not os.path.exists(directory):
        os.makedirs(directory)
    file_path = os.path.join(directory, f"{taskid}.xlsx")
    saved = workbook.save(file_path)
    print(f"file saved: {saved}")

    task_progress = {
        "status": "END",
        "taskid": self.request.id,
        "file": file_path,
    }
    async_to_sync(channel_layer.group_send)(
        self.request.id,  # Use the task ID as the channel group name
        {
            "type": "task_progress",
            "percentage_complete": task_progress,
        },
    )


def fetch_all_datas(queryset_values):
    data = []
    c = 1
    for each in queryset_values:
        ver = {
            "Serial No": c,
            "Review": str(each["text"]).title(),
            "Sentiment": each["sentiment"],
        }
        data.append(ver)
        c += 1
    return data


@shared_task(bind=True)
def download_data(self, queryset_values):
    print("______CELERY STARTS_______")
    # time.sleep(15)

    task_progress = {
        "status": 0,
        "taskid": self.request.id,
    }
    channel_layer = get_channel_layer()
    async_to_sync(channel_layer.group_send)(
        self.request.id,  # Use the task ID as the channel group name
        {"type": "task_progress", "percentage_complete": task_progress},
    )
    finaldatas = fetch_all_datas(queryset_values)
    write_to_excel_file(self, finaldatas)
    return "DONEEE**DOWNLOAD***********"


@shared_task(bind=True)
def do_sentiment_analysis(self, text):
    print("do_sentiment_analysis of ", text)
    selected_category = None
    category = None
    task_progress = {
        "status": "sentiment_updated",
        "taskid": {
            "taskid": self.request.id,
            "percentage": "10",
        },
    }
    channel_layer = get_channel_layer()
    async_to_sync(channel_layer.group_send)(
        self.request.id,  # Use the task ID as the channel group name
        {"type": "task_progress", "percentage_complete": task_progress},
    )

    # Allocate a pipeline for sentiment-analysis
    classifier = pipeline("sentiment-analysis")
    analysis = classifier(text)
    print(analysis)
    label = analysis[0].get("label", None)
    print("LABEL: ", label)

    sentiment = {
        "POSITIVE": "Positive",
        "NEUTRAL": "Neutral",
        "NEGATIVE": "Negative",
    }
    label = sentiment[label] if label in sentiment.keys() else "Other"
    print("label ", label)

    task_progress = {
        "status": "sentiment_updated",
        "taskid": {
            "taskid": self.request.id,
            "percentage": "50",
        },
    }
    channel_layer = get_channel_layer()
    async_to_sync(channel_layer.group_send)(
        self.request.id,  # Use the task ID as the channel group name
        {"type": "task_progress", "percentage_complete": task_progress},
    )

    # Allocate a pipeline for text-classification
    model_name = "M47Labs/english_news_classification_headlines"
    try:
        classifier2 = pipeline(task="text-classification", model=model_name)
        analysis2 = classifier2(text)
        print(analysis2)
        label_category = str(analysis2[0].get("label", None)).strip().upper()

        if Categories.objects.filter(name=label_category).exists():
            category = Categories.objects.get(name=label_category)
        else:
            category = Categories.objects.create(name=label_category)
    except:
        try:
            if label:
                if Categories.objects.filter(name=label.upper()).exists():
                    category = Categories.objects.get(name=label.upper())
                else:
                    category = Categories.objects.create(name=label.upper())

        except Exception as err:
            print("Exception:", err)

    print("--category name-", category.name)

    # # create sentence
    sentence = Sentences.objects.create(text=text, category=category)
    sentence.sentiment = label
    sentence.save()

    selected_category = sentence.category.name

    task_progress = {
        "status": "sentiment_updated",
        "taskid": {
            "taskid": self.request.id,
            "percentage": "100",
            "selected_category": selected_category,
        },
    }
    channel_layer = get_channel_layer()
    async_to_sync(channel_layer.group_send)(
        self.request.id,  # Use the task ID as the channel group name
        {"type": "task_progress", "percentage_complete": task_progress},
    )
